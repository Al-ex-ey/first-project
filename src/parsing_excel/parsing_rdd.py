import os
import re
import openpyxl
from pathlib import Path

BASE_DIR = Path(__file__).parent


rdd_pattern = re.compile(r"rdd\_\d\d\.\d\d\d\d\.xlsx$")

downloads_dir = BASE_DIR/"downloads"
files_dir = os.listdir(downloads_dir)

rdd_list = []

for file in files_dir:
    if  re.search(rdd_pattern, file):
        rdd_list.append(file)

if len(rdd_list) > 0:
    rdd_files = [os.path.join(downloads_dir, file) for file in rdd_list]
    rdd_files = [file for file in rdd_files if os.path.isfile(file)]
    rdd_file_path = max(rdd_files, key=os.path.getctime)
    rdd_file_name = rdd_file_path.split('\\')[-1]
    check_rdd_pattern = re.search(rdd_pattern, rdd_file_name)
    if check_rdd_pattern:            
        rdd_dir = rdd_file_path

book_rdd = openpyxl.load_workbook(filename=rdd_dir)

sheet_rdd = book_rdd.worksheets[0]

fio = []
for i in range(1, 10000):
    manager_cell = sheet_rdd.cell(row=i, column=1)
    rows_rdd = sheet_rdd.cell(row=i, column=1)
    if sheet_rdd.value == "Итого":
        break
    if manager_cell.value == None or manager_cell.alignment.indent != 0:
        continue
    fio.append(manager_cell.value)

for i in range(1, 10000):
    rows_rdd = sheet_rdd.cell(row=i, column=1)
    if sheet_rdd.value == "Чуквасов Алексей Юрьевич":
        break
    else:
        rows_rdd.delete_rows(i, amount=1)


# arendator_cell = sheet_arenda.cell(row=a, column=1)

# for a in range(1, 333):
#     if arendator_cell.value == "КОНТАКТЫ":
#         break
    
book_rdd.save(rdd_dir)


print(fio)
