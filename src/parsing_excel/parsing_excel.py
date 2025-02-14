import os
import re
import openpyxl
import logging
# import pandas as pd
import datetime as dt
import shutil
from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import RedirectResponse

# from openpyxl.styles import PatternFill
# from openpyxl.styles.differential import DifferentialStyle
# from openpyxl.formatting.rule import Rule
from openpyxl.styles import NamedStyle, Alignment, Font, Border, Side, PatternFill
from src.configs import configure_logging, settings
from src.constants import (
    AMOUNT_ROW,
    AMOUNT_A,
    AMOUNT_ROW_TOTAL,
    AMOUNT_A_TOTAL,
    BASE_DIR,
    ARENDA_AMOUNT_ROW,
    DEBIT_AMOUNT_ROW,
    DT_FORMAT,
    DATE_NOW,
)
from pathlib import Path
from tqdm import tqdm


configure_logging()

router = APIRouter()

now = dt.datetime.now()
date = now.strftime(DATE_NOW)

# data = sheet_arenda.values
# df = pd.DataFrame(data)
# df.to_excel('C:/Users/user/Desktop/Arenda_df.xlsx', index=False, header=False, sheet_name=f"{now.strftime(DATE_NOW)}_auto")


def parsing_excel(AMOUNT_ROW_TOTAL, AMOUNT_ROW, AMOUNT_A, AMOUNT_A_TOTAL, ARENDA_AMOUNT_ROW, DEBIT_AMOUNT_ROW):

    """ Загрузка файлов для обработкии и проверка на соответствие"""
    # try:
    debet_pattern = re.compile(r"debet\_\d\d\.\d\d\.\d\d\d\d\_\d\d\-\d\d\.xlsx$")
    rdd_pattern = re.compile(r"rdd\_\d\d\.\d\d\d\d\.xlsx$")
    downloads_dir = BASE_DIR/"downloads"
    files_dir = os.listdir(downloads_dir)
    debet_list = []
    rdd_list = []
    for file in files_dir:
        if  re.search(debet_pattern, file):
            debet_list.append(file)
        if  re.search(rdd_pattern, file):
            rdd_list.append(file)
        logging.info(f"==================== parsing_excel - формирование списка файлов дебеторки и цены на арнеду в дерикторнии прошло успешно! ====================")

    if len(debet_list) > 0:
        debet_files = [os.path.join(downloads_dir, file) for file in debet_list]
        debet_files = [file for file in debet_files if os.path.isfile(file)]
        debet_file_path = max(debet_files, key=os.path.getctime)
        debet_file_name = debet_file_path.split('\\')[-1]
        check_debet_pattern = re.search(debet_pattern, debet_file_name)
        if check_debet_pattern:
            debet_dir = debet_file_path
            logging.info(f"==================== parsing_excel - проверка наличия файла с дебиторкой по патерну прошла успешно! ====================")
        else:
            logging.error(f"========== Файл с дебеторской задолженностью иммет не верное название ==========\n")
            raise HTTPException(status_code=404, detail="Ошибка! Проверьте: имя файла, что файл существует, не поврежден и в нем есть страницы!")
    else:
        logging.error(f"========== Файл с дебеторской задолженностью в директории не обнаружен ==========\n")
        raise HTTPException(
    status_code=404, detail="Ошибка! Проверьте, что файл существует, не поврежден и в нем есть страницы!"
    )

    if len(rdd_list) > 0:
        rdd_files = [os.path.join(downloads_dir, file) for file in rdd_list]
        rdd_files = [file for file in rdd_files if os.path.isfile(file)]
        rdd_file_path = max(rdd_files, key=os.path.getctime)
        rdd_file_name = rdd_file_path.split('\\')[-1]
        check_rdd_pattern = re.search(rdd_pattern, rdd_file_name)
        if check_rdd_pattern:
            rdd_dir = rdd_file_path
            logging.info(f"==================== parsing_excel - проверка наличия файла с ценой по аренде по патерну прошла успешно! ====================")
        else:
            logging.error(f"========== Файл с реестром действующих договоров иммет не верное название ==========\n")
    else:
        logging.error(f"========== Файл с реестром действующих договоров в директории не обнаружен ==========\n")

    if settings.file_name in files_dir:
        arenda_dir = downloads_dir/settings.file_name
        logging.info(f"==================== parsing_excel - проверка наличия файла аренды прошла успешно! ====================")
    else:
        logging.error(f"========== Файл {settings.file_name} в директории не обнаружен ==========\n")
        raise HTTPException(
            status_code=404, detail="Ошибка! Проверьте, что файлы существуют, не поврежденв и в них есть страницы!"
        )

    # except Exception:
    #     logging.error(f"========== Что то пошло не так при выборе файлов ==========\n")
    #     raise HTTPException (status_code=404, detail="Ошибка доступа к файлам, попробуйте еще раз!")

    """ Получение рабочих страниц из файлов """
    try:
        book_arenda = openpyxl.load_workbook(filename=arenda_dir)
        book_debet = openpyxl.load_workbook(filename=debet_dir)
        book_rdd = openpyxl.load_workbook(filename=rdd_dir)
        logging.info(f"==================== parsing_excel - книги созданы успешно! ====================")
        sheet_arenda = book_arenda.worksheets[-1]
        sheet_debet = book_debet.worksheets[0]
        sheet_rdd = book_rdd.worksheets[0]
        logging.info(f"==================== parsing_excel - рабочие страницы созданы успешно! ====================")
    except (OSError, IOError):
        logging.error(f"Ошибка доступа к файлу или отсутствует сетевое соединение!\n")
        raise HTTPException(
            status_code=404, detail="Ошибка доступа к файлу или отсутствует сетевое соединение!"
        )
    except Exception:
        logging.error(f"Ошибка! Проверьте, что файл существует, не поврежден и в нем есть страницы!\n")
        raise HTTPException(
            status_code=404, detail="Ошибка! Проверьте, что файл существует, не поврежден и в нем есть страницы!"
        )

    lost_contracts = []
    # arendators_not_in_debet_list = []
    # arendator_debet_list = []
    # arendator_list = []
    source = book_arenda.worksheets[-1]
    new_sheet = book_arenda.copy_worksheet(source)
    new_sheet.title = f'{date}_mh'

    """ Подготовка файла с дебеторкой: удаление не используемых столбцов и строк со страницы """
    try:
        column_list = [1,3,4,5,6]
        for i in reversed(column_list):
            sheet_debet.delete_cols(i)
    except:
        logging.error(f"Ошибка при обработке файла! Проверьте налицие необходимыхданных данных на странице.\n")
        raise

    try:
        sheet_debet.delete_rows(1, amount=7)
        for i in range(1, 3300):
            rows_debit = sheet_debet.cell(row=i, column=1)
            if rows_debit.value == "Итого":
                sheet_debet.delete_rows(i, amount=2)
                break
    except:
        logging.error(f"Ошибка при обработке файла! Проверьте налицие необходимыхданных данных на странице.\n")
        raise
    logging.info(f"==================== parsing_excel - Подготовка файла с дебеторкой: удаление не используемых столбцов и строк со страницы прошла успешно! ====================")

    # """ Разметка новой страницы - выделение нужных столбцов желтым фоном и удаление данных"""
    # for a in range(1, ARENDA_AMOUNT_ROW):
    #     arendator_cell = sheet_arenda.cell(row=a, column=1)
    #     if arendator_cell.value == "КОНТАКТЫ":
    #         break
    #     AMOUNT_ROW_TOTAL = AMOUNT_ROW_TOTAL + 1
    #     if (arendator_cell.value == None) or (arendator_cell.font.bold == True):
    #         continue
        # new_sheet.cell(a, column=3, value='').fill = PatternFill(fill_type='solid', start_color='FFFFC0')
        # new_sheet.cell(a, column=4, value='').fill = PatternFill(fill_type='solid', start_color='FFFFC0')
        # new_sheet.cell(a, column=5, value='').fill = PatternFill(fill_type='solid', start_color='FFFFC0')

    # pattern = re.compile(r'\w+[\-|\.]?\w+')
    # if check_rdd_pattern:
    """ Подготовка файла с ценами: удаление не используемых столбцов и строк со страницы """
    merged_ranges = list(sheet_rdd.merged_cells.ranges)

    for merged_range in merged_ranges:
        sheet_rdd.unmerge_cells(str(merged_range))

    column_list = [2,3,4,5,6,8]
    for i in reversed(column_list):
        sheet_rdd.delete_cols(i)

    rows_to_delete = []

    for i in range(1, 5000):
        cell_value = sheet_rdd.cell(row=i, column=1).value
        if cell_value == "Чуквасов Алексей Юрьевич":
            break
        else:
            rows_to_delete.append(i)

    for row in reversed(rows_to_delete):
        sheet_rdd.delete_rows(row)

    contragents: list = []

    for i in range(1, 1500):
        a: dict = {}
        arendator_cell = sheet_rdd.cell(row=i, column=1)
        arendator_price_cell = arendator_cell.offset(column=1)
        if arendator_cell.value == "Итого":
            break
        if arendator_cell.value == None or arendator_cell.alignment.indent != 6:
            continue

        # a.append(arendator_cell.value)
        # a.append(arendator_price_cell.value)
        a[arendator_cell.value] = arendator_price_cell.value
        contragents.append(a)
    logging.info(f"==================== parsing_excel - Подготовка файла с ценами: удаление не используемых столбцов и строк со страницы прошла успешно! ====================")


    # for a in range(1, ARENDA_AMOUNT_ROW):
    #     if arendator_cell.value == "КОНТАКТЫ":
    #         break

    """ Цикл сравнения книги арендаторов с книгой дебеторки"""
    for i in tqdm(range(1, AMOUNT_ROW_TOTAL), desc="Wait a going process"):
        arendator_cell = sheet_arenda.cell(row=i, column=1)
        arendator = arendator_cell.value
        
        if arendator_cell.value == "КОНТАКТЫ":
            break
        
        AMOUNT_A_TOTAL = AMOUNT_A_TOTAL + 1
        
        if arendator_cell.value == None or arendator_cell.font.bold == True:
            continue
        
        new_sheet.cell(i, column=3, value='').fill = PatternFill(fill_type='solid', start_color='FFFFC0')
        new_sheet.cell(i, column=4, value='').fill = PatternFill(fill_type='solid', start_color='FFFFC0')
        new_sheet.cell(i, column=5, value='').fill = PatternFill(fill_type='solid', start_color='FFFFC0')
        new_sheet.cell(i, column=6).fill = PatternFill(fill_type='solid', start_color='FFFFC0')
        
        """ Выбор по арендатрам """
        for d in sheet_debet.iter_rows(min_row=10, max_row=DEBIT_AMOUNT_ROW, min_col=1, max_col=3):
            debet_arendator_cell = d[0]
            arendator_d = debet_arendator_cell.value
            if arendator_d == None or debet_arendator_cell.alignment.indent > 0:
                continue

            pattern = re.compile(r'[\"|\(|\)]')
            comparison = re.search(re.sub(pattern, '', arendator.lower()), re.sub(pattern, '', arendator_d.lower()))

            cell_arenda_contract = sheet_arenda.cell(row=i, column=2)
            if cell_arenda_contract.value == None:
                if arendator not in lost_contracts:
                    lost_contracts.append(arendator)
                continue
                        
            """ Выбор по договорам"""
            for r in range(1, 20):
                cell_debet_contract = debet_arendator_cell.offset(row=r)
                if comparison and (cell_arenda_contract.value == cell_debet_contract.value):
                    debet_cell = cell_debet_contract.offset(column=1)
                    credit_cell = cell_debet_contract.offset(column=2)
                    
                    arendator_price = next((d[arendator] for d in contragents if arendator in d), None)
                    if arendator_price == None:
                        arendator_price = 0
                                        
                    logging.info(f" {arendator} ----- {cell_arenda_contract.value} ----- {debet_cell.value} ----- {credit_cell.value}\n")
                    note_style = "Neutral"
                    if debet_cell.value:
                        style_bad = "Bad"
                        note_style = "Bad"
                    else:
                        style_bad = "Normal"
                        if debet_cell.value == None:
                            debet_cell.value = 0

                    if credit_cell.value:
                        style_good = "Good"
                        if credit_cell.value >= arendator_price:
                            note_style = "Good"
                    else:
                        style_good = "Normal"
                        if credit_cell.value == None:
                            credit_cell.value = 0
                    
                    bd = Side(style='thin', color="00000000")
                    new_sheet.cell(i, column=3, value=debet_cell.value).style = style_bad
                    new_sheet.cell(i, column=3).alignment = Alignment(horizontal="center", vertical="center")
                    new_sheet.cell(i, column=3).border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    new_sheet.cell(i, column=4, value=credit_cell.value).style = style_good
                    new_sheet.cell(i, column=4).alignment = Alignment(horizontal="center", vertical="center")
                    new_sheet.cell(i, column=4).border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    new_sheet.cell(i, column=5, value=arendator_price).style = "Normal"
                    new_sheet.cell(i, column=5).alignment = Alignment(horizontal="center", vertical="center")
                    new_sheet.cell(i, column=5).border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    new_sheet.cell(i, column=6).style = note_style
                    new_sheet.cell(i, column=6).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    new_sheet.cell(i, column=6).border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    AMOUNT_A = AMOUNT_A + 1
                    break

    book_arenda.save(arenda_dir)

    """ Проверка результата обработки: поиск необработанных позиций в созданом файле,
        поиск арендаторов без договоров, поиск необработанных позиций, подсчет общей задолженности и переплаты,
        формирование словаря с результатом обработки, подсчет обработанных строк в файле"""
    arendator_without_contract = []
    not_processed = {}

    path = BASE_DIR/"downloads"
    files_dir = os.listdir(path)
    if settings.file_name in files_dir:
       book_arenda = openpyxl.load_workbook(filename=arenda_dir)
    else:
        raise HTTPException(status_code=404, detail="File not found")

    sheet_arenda = book_arenda.worksheets[-1]
    result_table: list = []
    total_credit: float = 0
    total_debet: float = 0
    counter: int = 1

    for b in range(1, AMOUNT_A_TOTAL):
        key_counter: str = f"key_" + str(counter)           # счетчик позиции должника
        arendator_cell = sheet_arenda.cell(row=b, column=1) # Арендатор
        contract_cell = sheet_arenda.cell(row=b, column=2)  # Договор
        debet_cell = sheet_arenda.cell(row=b, column=3) # Дебет
        credit_cell = sheet_arenda.cell(row=b, column=4)    # Кредит
        savesum_cell = sheet_arenda.cell(row=b, column=5)    # Размер страховой суммы по договору
        note_cell = sheet_arenda.cell(row=b, column=6)    # Примечание
        email_cell = sheet_arenda.cell(row=b, column=9) # эл. почта
        phone_cell = sheet_arenda.cell(row=b, column=10) # Телефон
        contact_cell = sheet_arenda.cell(row=b, column=11) # Контактная информация

        if arendator_cell.value == "КОНТАКТЫ":  # достигнут конец таблицы
            break
        if arendator_cell.font.bold == True:
            ul = arendator_cell.value # ul - юридическое лицо
        if (arendator_cell.value == None) or (arendator_cell.font.bold == True):    # пропускаем пустые и жирные(заголовки) строки
            continue
        AMOUNT_ROW = AMOUNT_ROW + 1 # счетчик обработанных позиций
        if contract_cell.value == None:
            arendator_without_contract.append(arendator_cell.value)
            continue
        if debet_cell.fill == PatternFill(fill_type='solid', start_color='FFFFC0') or credit_cell.fill == PatternFill(fill_type='solid', start_color='FFFFC0'): # не обработанные строки - в талице выделены желтым
            not_processed[arendator_cell.value] = contract_cell.value
        if debet_cell.value == None:
            if credit_cell.value == None:
                continue
        if debet_cell.value > 0:
            rt: dict = {}
            if key_counter not in rt:
                rt[key_counter] = []
                counter = counter + 1
            rt[key_counter].append(arendator_cell.value)
            rt[key_counter].append(contract_cell.value)
            rt[key_counter].append(debet_cell.value)
            rt[key_counter].append(credit_cell.value)
            rt[key_counter].append(email_cell.value)
            rt[key_counter].append(phone_cell.value)
            rt[key_counter].append(savesum_cell.value)
            rt[key_counter].append(note_cell.value)
            rt[key_counter].append(contact_cell.value)
            rt[key_counter].append(ul)
            result_table.append(rt)
            total_debet = total_debet + debet_cell.value
        if credit_cell.value > 0:
            total_credit = total_credit + credit_cell.value


    # AMOUNT_ROW_TOTAL = AMOUNT_A + len(arendator_without_contract) + len(not_processed)
    amount_wrong_row = len(not_processed) + len(arendator_without_contract)
    if AMOUNT_ROW != AMOUNT_A + amount_wrong_row:
        color = 31
    else:
        color = 32
    print(f"\033[1;{color};40m ===== {AMOUNT_A} обработанных строк из {AMOUNT_ROW}  ===== в том числе {amount_wrong_row} строк(а) без обработки ===== \033[0;0m\n")
    print("Обратите внимание:\n")
    print(f"\033[1;31;40m ===== Необработанные позиции ===== {not_processed}  \033[0;0m\n")
    print(f"\033[1;31;40m ===== Арендаторы без договоров ===== {arendator_without_contract}  \033[0;0m\n")
    print(f"\033[1;31;40m ===== Обработанный файл с дебеторкой {arenda_dir} в директории {downloads_dir}  \033[0;0m\n")


    # print(f"\033[1;{color};40m ========== {AMOUNT_A} строк в выводе из {AMOUNT_ROW} (арендаторы в файле аренда) ========== \033[0;0m\n")
    # print(f"Аредаторы без договоров____{lost_contracts}\n")
    # print(f"Аредаторы отсутствующте в фале дебеторки - {arendators_not_in_debet_list}\n")
    # # print(f"Аредаторы с некорректными договорами: {incorrect_contracts}\n")
    logging.info(f"\033[1;{color};40m ===== {AMOUNT_A} обработанных строк из {AMOUNT_ROW}  ===== в том числе {amount_wrong_row} строк(а) без обработки ===== \033[0;0m\n")
    # logging.info(f"Аредаторы без договоров____{lost_contracts}\n")
    # logging.info(f"Аредаторы отсутствующте в фале дебеторки - {arendators_not_in_debet_list}\n")
    # # logging.info(f"Аредаторы с некорректными договорами: {incorrect_contracts}\n")
    logging.info(f"Обработанный файл с дебеторкой {arenda_dir} в директории {downloads_dir}\n")

    total_credit = int(total_credit)
    total_debet = int(total_debet)
    query_params = {
        "arendator_without_contract": arendator_without_contract,
        "not_processed": not_processed,
        "file_name": arenda_dir,
        "file_path": downloads_dir,
        "result_table": result_table,
        "total_debet": total_debet,
        "total_credit": total_credit,
    }

    return query_params



# if not 'style_normal' in book_arenda.named_styles:
#     style_normal = NamedStyle(name="style_normal")
#     style_normal.font = Font(name='Times New Roman', bold=False, size=11, color="00000000")
#     style_normal.alignment = Alignment(horizontal='center', vertical='center')
#     style_normal.fill = PatternFill(fgColor="00FFFFFF")
#     bd = Side(style='thin', color="00000000")
#     style_normal.border = Border(left=bd, top=bd, right=bd, bottom=bd)
# else:
#     book_arenda.named_styles(name="style_normal")


# book_arenda.add_named_style(style_normal)

# for index,cur_style in enumerate(book_arenda._named_styles):
#     if cur_style.name == 'my_new_style':
#         book_arenda._named_styles[index] = style_normal
#         style_normal.bind(book_arenda)
#         break
# else:
#     book_arenda.add_named_style(style_normal)



# bg_red = PatternFill(fill_type='solid', fgColor="FF0000")
# diff_style = DifferentialStyle(fill=bg_red)
# rule = Rule(type="expression", dxf=diff_style)
# rule.formula = ["$B3>0"]
# new_sheet.conditional_formatting.add("C3:C100", rule)

# if 'style_normal' in book_arenda.named_styles:
#     del book_arenda.named_styles.del  = style_normal



# title = f"{now.strftime(DATE_NOW)}_auto"
# book_arenda.create_sheet(title)
# source = book_arenda.worksheets[-1]
# new_sheet = book_arenda.copy_worksheet(source)
# with ExcelWriter(PATH, mode="a", if_sheet_exists="new") as writer:
#     book_arenda.
# sheet = source.get_sheet_by_name()
# sheet.title = title
# book_arenda.save(PATH)


# for row in sheet_debet.iter_rows(min_row=10, max_row=3000, min_col=1, max_col=3):
    # if row[0].value == None:
    #     continue