import re
import openpyxl
import logging
# import pandas as pd
import datetime as dt

from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.styles import NamedStyle, Alignment, Font, Border, Side
from configs import configure_logging
from constants import AMOUNT_RAW, AMOUNT_A, AMOUNT_RAW_TOTAL, AMOUNT_A_TOTAL, PATH_A, PATH_D, ARENDA_AMOUNT_ROW, DEBIT_AMOUNT_ROW, DT_FORMAT
from tqdm import tqdm

configure_logging()
now = dt.datetime.now()

try:
    book_arenda = openpyxl.load_workbook(filename=PATH_A)
    book_debet = openpyxl.load_workbook(filename=PATH_D)
except:
    logging.ERROR(f"___{now.strftime(DT_FORMAT)}___Something went wrong when open and reed files")
    raise

sheet_arenda = book_arenda.worksheets[-1]
sheet_debet = book_debet.worksheets[0]
source = book_arenda.worksheets[-1]
new_sheet = book_arenda.copy_worksheet(source)
lost_contracts = []
arendators_not_in_debet_list = []
arendator_debet_list = []
arendator_list = []

# data = sheet_arenda.values
# df = pd.DataFrame(data)
# df.to_excel('C:/Users/user/Desktop/Arenda_df.xlsx', index=False, header=False, sheet_name=f"{now.strftime(DATE_NOW)}_auto")

try:
    column_list = [1,3,4,5,6]
    for i in reversed(column_list):
        sheet_debet.delete_cols(i)
except:
    logging.ERROR(f"___{now.strftime(DT_FORMAT)}___Something went wrong when delete columns")
    raise

try:
    sheet_debet.delete_rows(1, amount=7)
    for i in range(1, 2900):
        rows_debit = sheet_debet.cell(row=i, column=1)
        if rows_debit.value == "Итого":
            sheet_debet.delete_rows(i, amount=2)
            break
except:
    logging.ERROR(f"___{now.strftime(DT_FORMAT)}___Something went wrong when delete rows")
    raise


def find_debit(arendator, i):
    global AMOUNT_A
    for d in sheet_debet.iter_rows(min_row=10, max_row=DEBIT_AMOUNT_ROW, min_col=1, max_col=3):
        debit_arendator_cell = d[0]
        if (debit_arendator_cell.value == None) or (debit_arendator_cell.alignment.indent > 0):
            continue

        comparison = re.search(arendator.lower(), debit_arendator_cell.value.lower())

        cell_arenda_contract = sheet_arenda.cell(row=i, column=2)
        if cell_arenda_contract.value == None:
            if arendator not in lost_contracts:
                lost_contracts.append(arendator)
            continue

        for contract in range(1, 20):
            cell_debet_contract = debit_arendator_cell.offset(row=contract)
            try:
                if comparison and (cell_arenda_contract.value == cell_debet_contract.value):
                    debet_amount = cell_debet_contract.offset(column=1)
                    credit_amount = cell_debet_contract.offset(column=2)
                    logging.info(f"___ {arendator} ----- {cell_arenda_contract.value} ----- {debet_amount.value} ----- {credit_amount.value}\n")
                    if not debet_amount.value:
                        style_bad = "Normal"
                    else:
                        style_bad = "Bad"
                    if credit_amount.value:
                        style_good = "Good"
                    else:
                        style_good = "Normal"
                    if debet_amount.value == None:
                        debet_amount.value = 0
                    if credit_amount.value == None:
                        credit_amount.value = 0
                    bd = Side(style='thin', color="00000000")
                    new_sheet.cell(i, column=3, value=debet_amount.value).style = style_bad
                    new_sheet.cell(i, column=3).alignment = Alignment(horizontal="center", vertical="center")
                    new_sheet.cell(i, column=3).border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    new_sheet.cell(i, column=4, value=credit_amount.value).style = style_good
                    new_sheet.cell(i, column=4).alignment = Alignment(horizontal="center", vertical="center")
                    new_sheet.cell(i, column=4).border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    AMOUNT_A = AMOUNT_A + 1
            except:
                raise


def parsing_excel():
    global AMOUNT_RAW
    for a in range(1, ARENDA_AMOUNT_ROW):
        arendator_cell = sheet_arenda.cell(row=a, column=1)
        if (arendator_cell.value == None) or (arendator_cell.font.bold == True) or (arendator_cell.value in arendator_list):
            continue
        arendator_list.append(arendator_cell.value)
        # logging.info(f"{arendator_cell.value}")

    for d in range(1, DEBIT_AMOUNT_ROW):
        debit_arendator_cell = sheet_debet.cell(row=d, column=1)
        if (debit_arendator_cell.value == None) or (debit_arendator_cell.alignment.indent > 0) or (debit_arendator_cell.value in arendator_debet_list):
            continue
        arendator_debet_list.append(debit_arendator_cell.value)
        # logging.info(f"{debit_arendator_cell.value}")

    # pattern = re.compile(r'\S+')
    # for i in arendator_list:
    #     # i = str(i)
    #     # print(i)
    #     arendator = pattern.findall(i)
    #     print(arendator)
    #     print(len(arendator))
    #     if len(arendator) < 2:
    #         arendator = pattern.findall(arendator[0])
    #         print(arendator)
    #     else:
    #         arendator = pattern.findall(arendator)[:1]
    #         print(arendator)
    #     for p in arendator_debet_list:
    #         # print(p)
    #         comparison = re.search(str(arendator).lower(), p.lower())
    #         if not comparison:
    #             arendators_not_in_debet_list.append(i)

    pattern = re.compile(r'\S+')
    for i in tqdm(range(1, 230), desc="Wait a going process"):
        arendator_cell = sheet_arenda.cell(row=i, column=1)
        if arendator_cell.value == None:
            continue
        AMOUNT_RAW = AMOUNT_RAW + 1
        if arendator_cell.font.bold == True:
            continue
        arendator_pattern = pattern.findall(arendator_cell.value)
        if len(arendator_pattern) < 2:
            arendator = pattern.findall(arendator_cell.value)[0]
        else:
            arendator = pattern.findall(arendator_cell.value)[0] + " " + pattern.findall(arendator_cell.value)[1]
        find_debit(arendator, i)


if __name__ == '__main__':
    parsing_excel()
    book_arenda.save(PATH_A)

    if AMOUNT_RAW != AMOUNT_RAW_TOTAL or AMOUNT_A != AMOUNT_A_TOTAL:
        color = 31
    else:
        color = 32
    print(f"\033[1;{color};40m ========== {AMOUNT_RAW} строк обработано из {AMOUNT_RAW_TOTAL} расчетных в файле аренда ========== {AMOUNT_A} строк в выводе из {AMOUNT_A_TOTAL} расчетных (арендаторы в файле аренда) ========== \033[0;0m\n")
    print(f"Аредаторы без договоров____{lost_contracts}\n")
    print(f"=========================================================================================================================")
    print(f"Аредатор отсутствует в дебиторке - {arendators_not_in_debet_list}\n")



# if not 'style_normal' in book_arenda.named_styles:
#     style_normal = NamedStyle(name="style_normal")
#     style_normal.font = Font(name='Times New Roman', bold=False, size=11, color="00000000")
#     style_normal.alignment = Alignment(horizontal='center', vertical='center')
#     style_normal.fill = PatternFill(fgColor="00FFFFFF")
#     bd = Side(style='thin', color="00000000")
#     style_normal.border = Border(left=bd, top=bd, right=bd, bottom=bd)
# else:
#     book_arenda.named_styles(name="style_normal")


# book_arenda.add_named_style(style_normal)

# for index,cur_style in enumerate(book_arenda._named_styles):
#     if cur_style.name == 'my_new_style':
#         book_arenda._named_styles[index] = style_normal
#         style_normal.bind(book_arenda)
#         break
# else:
#     book_arenda.add_named_style(style_normal)



# bg_red = PatternFill(fill_type='solid', fgColor="FF0000")
# diff_style = DifferentialStyle(fill=bg_red)
# rule = Rule(type="expression", dxf=diff_style)
# rule.formula = ["$B3>0"]
# new_sheet.conditional_formatting.add("C3:C100", rule)

# if 'style_normal' in book_arenda.named_styles:
#     del book_arenda.named_styles.del  = style_normal



# title = f"{now.strftime(DATE_NOW)}_auto"
# book_arenda.create_sheet(title)
# source = book_arenda.worksheets[-1]
# new_sheet = book_arenda.copy_worksheet(source)
# with ExcelWriter(PATH, mode="a", if_sheet_exists="new") as writer:
#     book_arenda.
# sheet = source.get_sheet_by_name()
# sheet.title = title
# book_arenda.save(PATH)


# for row in sheet_debet.iter_rows(min_row=10, max_row=3000, min_col=1, max_col=3):
    # if row[0].value == None:
    #     continue